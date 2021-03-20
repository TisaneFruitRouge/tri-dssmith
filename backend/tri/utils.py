import string
from openpyxl import Workbook, load_workbook

def number_to_col(n: int):
		
	alphabet = string.ascii_uppercase

	if n>26: return alphabet[n//26]+alphabet[n%26]
	else   : return alphabet[n]


def insert_fichier(path: str, data):
	wb = load_workbook(filename=path)
	ws = wb.active

	data_items = data.items()

	data_keys = list(data)

	taille_data = len(data_items)

	list_data = [[0 for x in range(2)] for y in range(taille_data)] # Création de la liste qui contiendra les données

	for i in range(len(data_items)):
		list_data[i][0] = data_keys[i]
		list_data[i][1] = data[data_keys[i]]

	print(list_data)	

	for row in range(len(list_data)):
		for col in range(len(list_data[row])):
			n = number_to_col(col)
			new_worksheet.write(row, n, list_data[row][col])

	new_workbook.close()



def creer_fichier(path: str):
	new_workbook = Workbook(path)
	new_workbook.save(path)

'''def insert_fichier(path: str):
	new_workbook = xlsxwriter.Workbook(new_excel_path)
	new_worksheet = new_workbook.add_worksheet()
			

	data_items = data.items()

	data_keys = list(data)

	taille_data = len(data_items)
			


	list_data = [[0 for x in range(2)] for y in range(taille_data)] 

	for i in range(len(data_items)):
		list_data[i][0] = data_keys[i]
		list_data[i][1] = data[data_keys[i]]

	print(list_data)	

	for row in range(len(list_data)):
		for col in range(len(list_data[row])):
			new_worksheet.write(col, row, list_data[row][col])

	new_workbook.close()'''